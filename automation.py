import time
import random
from openpyxl.styles import PatternFill
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from selenium.webdriver.chrome.options import Options

# Chrome options setup
options = Options()
options.add_experimental_option("detach", True)
options.add_argument("--disable-blink-features=AutomationControlled")
options.add_argument("window_size=1280,800")
options.add_argument("--disable-popup-blocking")

# Initialize WebDriver
driver = webdriver.Chrome(options=options)

# Open the login page
driver.get('https://winmarkseller.com/products')

def find_element_by_xpath(xpath, wait_time=60, condition=EC.presence_of_element_located):
    """Helper function to wait for an element by XPath with specified wait time and condition."""
    return WebDriverWait(driver, wait_time).until(condition((By.XPATH, xpath)))

try:
    # Login process - presence check for login fields
    email_elem = find_element_by_xpath('//*[@id="email"]', condition=EC.presence_of_element_located)
    password_elem = find_element_by_xpath('//*[@id="password"]', condition=EC.presence_of_element_located)
    next_button = find_element_by_xpath('//*[@id="next"]', condition=EC.element_to_be_clickable)

    email_elem.send_keys('leewasser72@gmail.com')
    password_elem.send_keys('Peacock123!')
    next_button.click()


    # Wait for login to complete and ensure the next page is loaded
    find_element_by_xpath('//*[@id="root"]/div/div[1]/div[2]/div/div[1]/div/a[5]/h2')

    # Click on the next page element - from now on, make elements clickable
    product_section = find_element_by_xpath('//*[@id="root"]/div/div[1]/div[2]/div/div[1]/div/a[5]/h2',
                                            condition=EC.element_to_be_clickable)
    product_section.click()

    # Wait for the page with product list to load
    find_element_by_xpath('//*[@id=":r5:"]', condition=EC.element_to_be_clickable)

    # Load the Excel file containing product IDs
    wb = load_workbook('skus to fix ecom.xlsx')
    sheet = wb.active

    # Loop through the product IDs
    for row in range(1, sheet.max_row + 1):
        product_id = sheet.cell(row=row, column=1).value  # Get product ID from the first column

        try:
            # Input the product ID into the field
            product_id_field = find_element_by_xpath('/html/body/div[1]/div/div[2]/div/div[2]/div[1]/div/span/div[1]/input',
                                                     condition=EC.element_to_be_clickable)
            product_id_field.clear()
            product_id_field.send_keys(str(product_id))

            # Wait for the link to be clickable and click on it
            edit_button = find_element_by_xpath('/html/body/div[1]/div/div[2]/div/div[3]/table/tbody/tr/td[2]/div/div[2]/a',
                                                wait_time=40, condition=EC.element_to_be_clickable)
            edit_button.click()

            # Scroll down to check checkboxes
            ActionChains(driver).move_to_element(find_element_by_xpath('//*[@id="root"]/div/div[2]/div[3]/div[2]/form/div[6]/div[1]/label',
                                                                      condition=EC.element_to_be_clickable)).perform()

            # Locate the checkboxes
            checkbox1 = find_element_by_xpath('//*[@id="root"]/div/div[2]/div[3]/div[2]/form/div[6]/div[1]/label',
                                              condition=EC.element_to_be_clickable)
            checkbox2 = find_element_by_xpath('//*[@id="root"]/div/div[2]/div[3]/div[2]/form/div[6]/div[2]/label',
                                              condition=EC.element_to_be_clickable)

            # Check the class names to determine the checkbox state
            if checkbox1.get_attribute("class") == "styled__StyledCheckbox-sc-s1u0st-3 dJsvSf":  # Unchecked class
                checkbox1.click()  # Click to check the box
                WebDriverWait(driver, 1)  # Wait a bit before proceeding

            if checkbox2.get_attribute("class") == "styled__StyledCheckbox-sc-s1u0st-3 dJsvSf":  # Unchecked class
                checkbox2.click()  # Click to check the box

            # Scroll further down to input weight
            ActionChains(driver).move_to_element(find_element_by_xpath('//*[@id="weight"]',
                                                                      condition=EC.element_to_be_clickable)).perform()
            weight_field = find_element_by_xpath('//*[@id="weight"]', condition=EC.element_to_be_clickable)
            weight_field.clear()
            weight_field.send_keys(str(random.randint(5, 7)))

            # Scroll back up to the save button
            ActionChains(driver).move_to_element(
                find_element_by_xpath('/html/body/div[1]/div/div[2]/div[1]/div/div[2]/button',
                                      condition=EC.element_to_be_clickable)).perform()

            # Click on the Save Changes button
            save_button = find_element_by_xpath('/html/body/div[1]/div/div[2]/div[1]/div/div[2]/button',
                                                condition=EC.element_to_be_clickable)
            print(save_button.text)
            save_button.click()

            time.sleep(5)
            # Update the color of the product ID in the Excel sheet (highlight cell)
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            sheet.cell(row=row, column=1).fill = yellow_fill  # Highlight the cell in yellow

            # Save the Excel file after updating the color
            wb.save('skus to fix ecom.xlsx')

            # Go back to the products page and wait for elements to load again
            driver.get('https://winmarkseller.com/products')
            # Re-find the product list to make sure the page is fully loaded
            find_element_by_xpath('/html/body/div[1]/div/div[2]/div/div[2]/div[1]/div/span/div[1]/input', condition=EC.presence_of_element_located)

        except Exception as e:
            print(f"Error processing product ID {product_id}: {e}")
    print("All numbers are processed.")
except Exception as e:
    print(f"An error occurred during login or processing: {e}")
finally:
    driver.quit()
